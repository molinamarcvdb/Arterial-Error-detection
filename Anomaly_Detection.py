#Common imports
import os
import slicer
import vtk
import numpy as np
import JupyterNotebooksLib as slicernb
import skimage 
import pandas as pd
import SimpleITK as sitk
import nibabel as nib
from time import time
import matplotlib.pyplot as plt
import scipy
from scipy.spatial import distance
import networkx as nx
import nibabel as nib
from nibabel.affines import apply_affine
import openpyxl
from openpyxl import workbook
import math
import argparse

# Change smoothing factor - 12494908, - 
#slicer.util.pip_install('openpyxl')

def main(args):
    casePath = args.casePath

def make_parser():
    parser = argparse.ArgumentParser(description = 'Wrapping Anomaly_Detction_ALgorithm')
    parser.add_argument('-casePath', type=str, required=True, help = 'Path to directory containing the nifti, the labeled graph, the segments array and the labeled Excel')
    return parser

if __name__ == '__main__':
    parser = make_parser()
    args = parser.parse_args()

main(args)
casePath = args.casePath

# Defining necessary functions
def bbox_3D(img):

    LR = np.any(img, axis=(0, 1))
    PA = np.any(img, axis=(0, 2))
    IS = np.any(img, axis=(1, 2))

    minLR, maxLR = np.where(LR)[0][[0, -1]]
    minPA, maxPA = np.where(PA)[0][[0, -1]]
    minIS, maxIS = np.where(IS)[0][[0, -1]]

    return minLR, maxLR, minPA, maxPA, minIS, maxIS
def robustEndPointDetection(endpoint, segmentation, aff, n=10):
    ''' Relocates automatically detected endpoints to the center of mass of the closest component
    inside a local region around the endpoint (defined by n).

    Takes the endpoint position, converts it to voxel coordinates with the affine matrix, then defines a region  
    of (2 * n) ^ 3 voxels centered around the endpoint. Then components inside the local region are treated 
    as separate objects. The minimum distance from theese objects to the endpoint is computed, and from 
    these, the object with the smallest distance to the endpoint is chosen to compute the centroid, which
    is converted back to RAS with the affine matrix.

    Arguments:
        - endpoint <np.array>: position of the endpoint in RAS coordinates.
        - segmentation <np.array>: numpy array corresponding to the croppedVolumeNode.
        - aff <np.array>: affine matrix corresponding ot he nifti file.
        - n <int>: defines size of the region around the endpoint that is analyzed for this method.

    Returns:
        - newEndpoint <np.array>: new position of the endpoint.

    '''

    from skimage.measure import regionprops, label
    from scipy import ndimage

    # Compute RAS coordinates with affine matrix
    R0, A0, S0 = np.round(np.matmul(np.linalg.inv(aff), np.append(endpoint, 1.0))[:3]).astype(int)
    
    # Mask the segmentation (Only region of interest)
    maskedSegmentation = segmentation[np.max([0, S0 - n]): np.min([segmentation.shape[0], S0 + n]), 
                                      np.max([0, A0 - n]): np.min([segmentation.shape[1], A0 + n]),
                                      np.max([0, R0 - n]): np.min([segmentation.shape[2], R0 + n])]
    
    # Divide into different connected components
    labelMask = label(maskedSegmentation)
    
    labels = np.sort(np.unique(labelMask))
    labels = np.delete(labels, np.where([labels == 0]))
    
    labelMaskOneHot = np.zeros([len(labels), labelMask.shape[0], labelMask.shape[1], labelMask.shape[2]], dtype=np.uint8)
    for idx, label in enumerate(labels):
        labelMaskOneHot[idx][labelMask == label] = 1
        
    invertedLabelMaskOneHot = np.ones_like(labelMaskOneHot) - labelMaskOneHot
    
    # Get distance transform for each and get only closest component
    distanceLabels = np.empty_like(labels, dtype=np.float)
    for idx in range(len(labels)):
        distanceLabels[idx] = ndimage.distance_transform_edt(invertedLabelMaskOneHot[idx])[invertedLabelMaskOneHot.shape[1] // 2][invertedLabelMaskOneHot.shape[2] // 2][invertedLabelMaskOneHot.shape[3] // 2]

    mask = np.zeros_like(segmentation)
    mask[np.max([0, S0 - n]): np.min([segmentation.shape[0], S0 + n]), 
         np.max([0, A0 - n]): np.min([segmentation.shape[1], A0 + n]),
         np.max([0, R0 - n]): np.min([segmentation.shape[2], R0 + n])] = labelMaskOneHot[np.argmin(distanceLabels)]
    
    # Get the centroid of the foregroud region
    properties = regionprops(mask.astype(np.int), mask.astype(np.int))
    centerOfMass = np.array(properties[0].centroid)[[2, 1, 0]]
    
    # Return the new position of the endpoint in RAS coordinates
    return np.matmul(aff, np.append(centerOfMass, 1.0))[:3]


#############################Mac####################################################################
# Define caseID and path to directory 
caseDir = os.path.abspath(os.path.dirname(casePath))
caseId = f"{casePath[-15:]}"

# Obtain affine matrix 
aff = np.asarray(nib.load(os.path.join(caseDir, caseId)).affine)

# Load volume and associate to node
slicer.util.loadLabelVolume(casePath)
masterVolumeNode = getNode(os.path.basename(casePath[:-7])) # caseId is the file name without the extension (.nii.gz)

# Get initial time
start0 = time()

# Obtain NParray
Volumearray = slicer.util.arrayFromVolume(masterVolumeNode)
Volumearray.shape, aff 

# Set to 0 the voxels in the upper 20% of the bounding box
maskedVolumeArray = slicer.util.arrayFromVolume(masterVolumeNode)
_, _, _, _, minIS, maxIS = bbox_3D(maskedVolumeArray)
maskedVolumeArray[int(np.round((maxIS - minIS) * 0.85)):] = 0
slicer.util.updateVolumeFromArray(masterVolumeNode, maskedVolumeArray)

### Pre-Processing ###

# Create segmentation
segmentationNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLSegmentationNode")
segmentationNode.CreateDefaultDisplayNodes() # only needed for display

# Create segment editor to get access to effects
segmentEditorWidget = slicer.qMRMLSegmentEditorWidget()
segmentEditorWidget.setMRMLScene(slicer.mrmlScene)
segmentEditorNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLSegmentEditorNode")
segmentEditorNode.SetOverwriteMode(2) # Allow overlapping between different segments
segmentEditorWidget.setMRMLSegmentEditorNode(segmentEditorNode)

# Associate segmentationNode to segmentEditorWidget and add an empty segment
segmentationNode.SetReferenceImageGeometryParameterFromVolumeNode(masterVolumeNode)
addedSegmentID = segmentationNode.GetSegmentation().AddEmptySegment("Segmentation")
segmentEditorWidget.setSegmentationNode(segmentationNode)
segmentEditorWidget.setMasterVolumeNode(masterVolumeNode)

# Thresholding
segmentEditorWidget.setActiveEffectByName("Threshold")
effect = segmentEditorWidget.activeEffect()
effect.setParameter("MinimumThreshold","1")
effect.setParameter("MaximumThreshold","1")
effect.self().onApply()

# Smoothing
segmentEditorWidget.setActiveEffectByName("Smoothing")
effect = segmentEditorWidget.activeEffect()
effect.setParameter("SmoothingMethod", "JOINT_TAUBIN")
effect.setParameter("JointTaubinSmoothingFactor", 0.25)
effect.self().onApply()

# Remove MinimumSize segments
segmentEditorWidget.setActiveEffectByName("Islands")
effect = segmentEditorWidget.activeEffect()
effect.setParameter('Operation', "REMOVE_SMALL_ISLANDS")
effect.setParameter('MinimumSize', '1000')
effect.self().onApply()

# Split segments
segmentEditorWidget.setActiveEffectByName("Islands")
effect = segmentEditorWidget.activeEffect()
effect.setParameter('Operation', "SPLIT_ISLANDS_TO_SEGMENTS")
effect.self().onApply()

# Set view layout to desired one. Use this whenever you want to print 
slicernb.setViewLayout("OneUp3D")
slicernb.ViewDisplay()

# Make segmentation results visible in 3D
segmentationNode.CreateClosedSurfaceRepresentation()
start1 = time()
print(f"Data preparation took {start1 - start0} s")

### Endpoint Detection ###

# Clean up
segmentEditorWidget = None
slicer.mrmlScene.RemoveNode(segmentEditorNode)
listaEP = []
print("Beginning centerline extraction. Total number of segments:", segmentationNode.GetSegmentation().GetNumberOfSegments())
print()

GeneralEPNode = slicer.mrmlScene.AddNewNodeByClass('vtkMRMLMarkupsFiducialNode')
GeneralEPNode.SetName("GeneralEP")

for segmentId in range(segmentationNode.GetSegmentation().GetNumberOfSegments()):
    start40 = time()
    print("Segment", segmentId)
     
    # Set up extract centerline widget
    extractCenterlineWidget = None
    parameterNode = None
    # Instance Extract Centerline Widget
    extractCenterlineWidget = slicer.modules.extractcenterline.widgetRepresentation().self()
    # Set up parameter node
    parameterNode = slicer.mrmlScene.GetSingletonNode("ExtractCenterline", "vtkMRMLScriptedModuleNode")
    extractCenterlineWidget.setParameterNode(parameterNode)
    extractCenterlineWidget.setup()

    # Update from GUI to get segmentationNode as inputSurfaceNode
    extractCenterlineWidget.updateParameterNodeFromGUI()
    # Set network node reference to new empty node
    extractCenterlineWidget._parameterNode.SetNodeReferenceID("InputSurface", segmentationNode.GetID())
    extractCenterlineWidget.ui.inputSegmentSelectorWidget.setCurrentSegmentID(segmentationNode.GetSegmentation().GetNthSegmentID(segmentId))

    print("Automatic endpoint extraction...")
    # Autodetect endpoints
    extractCenterlineWidget.onAutoDetectEndPoints()
    extractCenterlineWidget.updateGUIFromParameterNode()
    

    start41 = time()
    print(f"Automatic endpoint detection took {start41 - start40} s")

    print("Relocating endpoints to center of mass of local closest object...")
    # Get volume node array from segmentation node
    labelmapVolumeNode = slicer.mrmlScene.AddNewNodeByClass('vtkMRMLLabelMapVolumeNode')
    slicer.modules.segmentations.logic().ExportAllSegmentsToLabelmapNode(segmentationNode, labelmapVolumeNode)
    segmentationArray = slicer.util.arrayFromVolume(labelmapVolumeNode)

    # Get affine matrix from segmentation labelMapVolumeNode
    vtkAff = vtk.vtkMatrix4x4()
    aff = np.eye(4)
    labelmapVolumeNode.GetIJKToRASMatrix(vtkAff)
    vtkAff.DeepCopy(aff.ravel(), vtkAff)

    # Get endpoints node
    endpointsNode = slicer.util.getNode(extractCenterlineWidget._parameterNode.GetNodeReferenceID("EndPoints"))
    NodeEP1 = slicer.mrmlScene.AddNewNodeByClass('vtkMRMLMarkupsFiducialNode')
    NodeEP1.CopyContent(endpointsNode)
    # Relocate endpoints for robust centerline extraction 
    
    for idx in range(endpointsNode.GetNumberOfFiducials()):
        endpoint = np.array(endpointsNode.GetCurvePoints().GetPoint(idx))
        newEndpoint = robustEndPointDetection(endpoint, segmentationArray, aff) # Center of mass of closest component method
        listaEP.append(newEndpoint)
        GeneralEPNode.AddFiducial(newEndpoint[0],
                                  newEndpoint[1],
                                  newEndpoint[2])
        endpointsNode.SetNthFiducialPosition(idx, newEndpoint[0],
                                                  newEndpoint[1],
                                                  newEndpoint[2])
    # Remove superposed endpoints 
    for idx in range(len(listaEP)):
        if distance.euclidean(listaEP[idx-1], listaEP[idx]) < 0.3 and distance.euclidean(listaEP[idx-2], listaEP[idx]) < 0.3:
            GeneralEPNode.RemoveMarkup(idx)
#     start42 = time()

########################
### Minimum Distance ###
########################

# Get Endpoints
listaEP = []
for idx in range(GeneralEPNode.GetNumberOfFiducials()):
    listaEP.append(GeneralEPNode.GetCurvePoints().GetPoint(idx))
# Get nº of EP    
nEP = len(listaEP)
# Obtein height value
H = []
for idx in range(len(listaEP)):
    H.append(listaEP[idx][2])
H
maxH = max(H)
minH = min(H)
Height = int(maxH - minH)

# Eliminate detections on the upper 5% 

for idx in range(len(listaEP)):
    if 0.95 * Height < listaEP[idx][2]-minH:
        listaEP[idx] = np.asanyarray((0, 0, 0))
    
# Considering the conflictive points depending on the distance lying btw 2 points. If it is lower than a certain valaue (i.e 20 or 25)
# Then we compute the middistance and the midpoint so that we obtain the central point of the bbox which is gonna be saved in NIFTI format

listEP = listaEP
#Distance btw each endpoint
dist3D = np.empty((len(listEP),len(listEP)))
indexX= []
indexY= []

for idx in range(len(listEP)):
    for idy in range(len(listEP)):
        Dist = distance.euclidean(listEP[idx], listEP[idy])
        dist3D[idx, idy] = Dist
        if Dist<25 and Dist>0:
                indexX.append(idx)
                indexY.append(idy)
        
df_distances = pd.DataFrame(dist3D)
Index = np.asarray(list(zip(indexX, indexY)))

# Save the ID's of each pair of endpoints
EP=[]
for idx in range(len(Index)):# Delete duplicated detections
    for idy in Index[idx]:
        if idy<Index[idx][1]:
            EP.append(Index[idx])         

            
# Middistance of each pair of endpoints
x=[]
y=[]
z=[] 
for idx in range(len(EP)):
    x.append((listEP[EP[idx][0]][0]+listEP[EP[idx][1]][0])/2)
    y.append((listEP[EP[idx][0]][1]+listEP[EP[idx][1]][1])/2)
    z.append((listEP[EP[idx][0]][2]+listEP[EP[idx][1]][2])/2)
    
Centerpoints=np.asarray(list(zip(x,y,z)))
# Get coordinates of each pair of endpoints
COORD1 = []
COORD2 = []
EPCOORD = []
for idx in range(len(EP)):
    COORD1.append(listEP[EP[idx][0]])
    COORD2.append(listEP[EP[idx][1]])
print("The ID's of the conflictive endpoints are:")
print((EP, len(EP)))


################################################
### Load labeled graph and SegmentsArray.npy ### 
################################################


# Load graph
G = nx.read_gpickle(os.path.join(caseDir, "graph_label.pickle"))
# Get cellId to segmentId dict
cellIdToVesselType = {}
for n0, n1 in G.edges:
    cellIdToVesselType[G[n0][n1]["CellID"]] = G[n0][n1]["edgetype"]
# {Dict CellID : Edgetype}
print()
print("We can asses the vessel type of each segment in SegementsArray following:")
print(cellIdToVesselType)
# Create a dictionary with EdgeTypes
edgeTypesDict = {
    0: "other",
    1: "AA",
    2: "BT",
    3: "RCCA",
    4: "LCCA",
    5: "RSA",
    6: "LSA",
    7: "RVA",
    8: "LVA",
    9: "RICA",
    10: "LICA",
    11: "RECA",
    12: "LECA",
    13: "BA",
    14: "AA+BT",
    15: "RVA+LVA"}

# Transform IJK to RAS coordinates
segmentsArray = np.load(os.path.join(caseDir, "segmentsArray.npy"), allow_pickle=True)
segmentsArrayAff = np.ndarray([len(segmentsArray)], dtype = object)
aff = np.asarray(nib.load(os.path.join(caseDir, caseId)).affine)
for idx in range(len(segmentsArray)):
    segmentsArrayAff[idx] = np.ndarray([len(segmentsArray[idx][0]), 3])
    for idx2 in range(len(segmentsArray[idx][0])):
        segmentsArrayAff[idx][idx2] = np.matmul(aff, np.append(segmentsArray[idx][0][idx2], 1.0))[:3]

        
##########################################
### Vertebral Arthery Origin Detection ###
##########################################      
        
# But still a big amount of the verterbal arthery origins remains being missdetected, so that we will develop a simple algorithm 
# which detects if the L/RSA & L/RVA segments are in contact. To do this we'll need to get the closest points of both segments. 
RSA =  []  # Type 5
LVA =  []  # Type 8
RVA =  []   # Type 7
LSA =  []  # Type 6
for idx in range(len(cellIdToVesselType)):
    if cellIdToVesselType[idx] == 7: #RVA
        RVA.append(np.asanyarray(segmentsArrayAff[idx]))
    if cellIdToVesselType[idx] == 5: #RSA
        RSA.append(segmentsArrayAff[idx])
    if cellIdToVesselType[idx] == 8: #LVA
        LVA.append(segmentsArrayAff[idx])
    if cellIdToVesselType[idx] == 6: #LSA
        LSA.append(segmentsArrayAff[idx])
# flatten - lists of lists
flat_RVA = [item for sublist in RVA for item in sublist]
flat_LVA = [item for sublist in LVA for item in sublist]
flat_RSA = [item for sublist in RSA for item in sublist]
flat_LSA = [item for sublist in LSA for item in sublist]

# Obtain the points of the L/RVA segments with the minimum S coordinate (RAS).

if len(flat_LVA) > 0:
    LVA_height = []
    for idx in range(len(flat_LVA)): 
        LVA_height.append(flat_LVA[idx][2])
    Min_LVA_idx = np.argmin(LVA_height)
    Min_LVA_Coord = flat_LVA[Min_LVA_idx]
    # Get the distance form all points of the L/RSA to the obtained coordinate of each vertebral arthery segment.
    #Get those (2) endpoints which are closer to the Min_L/RVA_Coord
    Distance_LSA = np.empty((len(flat_LSA),1))

    for idx in range(len(flat_LSA)):
        Distance_LSA[idx] = distance.euclidean(Min_LVA_Coord, flat_LSA[idx])
    Min_LSA_idx = np.argmin(Distance_LSA)
    Min_LSA_Coord = flat_LSA[Min_LSA_idx]

    if distance.euclidean(Min_LSA_Coord, Min_LVA_Coord) > 0.5:
        COORD1.append(Min_LSA_Coord)
        COORD2.append(Min_LVA_Coord)
        # Add L/RVA centerpoint to the centerpoints array
        x = (Min_LSA_Coord[0]+Min_LVA_Coord[0])/2
        y = (Min_LSA_Coord[1]+Min_LVA_Coord[1])/2
        z = (Min_LSA_Coord[2]+Min_LVA_Coord[2])/2
        Centerpoint_LVA = np.array([x,y,z])
        if len(Centerpoints) > 0:
            Centerpoints = np.append(Centerpoints, [Centerpoint_LVA], axis = 0)
        else:
            Centerpoints = np.array([Centerpoint_LVA])
        print()
        print(f"There's a dicontinuity in the LVA origin between:")
        print(f"{[Min_LSA_Coord, Min_LVA_Coord]}, the associated centerpoint has been appended to Centerpoints Array")

if len(flat_RVA) > 0:
    RVA_height = []
    for idx in range(len(flat_RVA)): 
        RVA_height.append(flat_RVA[idx][2])
    Min_RVA_idx = np.argmin(RVA_height)
    Min_RVA_Coord = flat_RVA[Min_RVA_idx]

    # Get the distance form all points of the L/RSA to the obtained coordinate of each vertebral arthery segment.
    #Get those (2) endpoints which are closer to the Min_L/RVA_Coord
    Distance_RSA = np.empty((len(flat_RSA),1))

    for idx in range(len(flat_RSA)):
        Distance_RSA[idx] = distance.euclidean(Min_RVA_Coord, flat_RSA[idx])
    Min_RSA_idx = np.argmin(Distance_RSA)
    Min_RSA_Coord = flat_RSA[Min_RSA_idx]


    if distance.euclidean(Min_RSA_Coord, Min_RVA_Coord) > 0.5:
        COORD1.append(Min_RSA_Coord)
        COORD2.append(Min_RVA_Coord)
        # Add L/RVA centerpoint to the centerpoints array
        x = (Min_RSA_Coord[0]+Min_RVA_Coord[0])/2
        y = (Min_RSA_Coord[1]+Min_RVA_Coord[1])/2
        z = (Min_RSA_Coord[2]+Min_RVA_Coord[2])/2
        Centerpoint_RVA = np.array([x,y,z])
        if len(Centerpoints) > 0:
            Centerpoints = np.append(Centerpoints, [Centerpoint_RVA], axis = 0)
        else:
            Centerpoints = np.array([Centerpoint_RVA])
        print()
        print(f"There's a dicontinuity in the RVA origin between:")
        print(f"{[Min_RSA_Coord, Min_RVA_Coord]}, the associated centerpoint has been appended to Centerpoints Array")

##########################################
### Vertebral-Basilar Arthery Detection ###
##########################################      
        
# But still a big amount of the verterbal arthery origins remains being missdetected, so that we will develop a simple algorithm 
# which detects if the L/RSA & L/RVA segments are in contact. To do this we'll need to get the closest points of both segments. 
BA =  []  # Type 13
LVA =  []  # Type 8
RVA =  []   # Type 7

for idx in range(len(cellIdToVesselType)):
    if cellIdToVesselType[idx] == 13: #BA
        BA.append(np.asanyarray(segmentsArrayAff[idx]))
    if cellIdToVesselType[idx] == 7: #RVA
        RVA.append(segmentsArrayAff[idx])
    if cellIdToVesselType[idx] == 8: #LVA
        LVA.append(segmentsArrayAff[idx])
    
# flatten - lists of lists
flat_RVA = [item for sublist in RVA for item in sublist]
flat_LVA = [item for sublist in LVA for item in sublist]
flat_BA = [item for sublist in BA for item in sublist]


# Obtain the points of the L/RVA segments with the maximum S coordinate (RAS).
Max_LVA_Coord = None
Max_RVA_Coord = None
Min_BA_Coord = None

if len(flat_LVA) > 0:
    LVA_height = []
    for idx in range(len(flat_LVA)): 
        LVA_height.append(flat_LVA[idx][2])
    Max_LVA_idx = np.argmax(LVA_height)
    Max_LVA_Coord = flat_LVA[Max_LVA_idx]

if len(flat_RVA) > 0:
    RVA_height = []
    for idx in range(len(flat_RVA)): 
        RVA_height.append(flat_RVA[idx][2])
    Max_RVA_idx = np.argmax(RVA_height)
    Max_RVA_Coord = flat_RVA[Max_RVA_idx]

    # Get the distance form all points of the L/RSA to the obtained coordinate of each vertebral arthery segment.
    #Get those (2) endpoints which are closer to the Min_L/RVA_Coord
if len(flat_BA) > 0 :
    BA_height = []
    for idx in range(len(flat_BA)): 
        BA_height.append(flat_BA[idx][2])
    Min_BA_idx = np.argmin(BA_height)
    Min_BA_Coord = flat_BA[Min_BA_idx] 

if len(flat_BA) > 0 and len(flat_LVA) > 0:
    if Max_RVA_Coord is not None and Max_LVA_Coord is not None:
        if distance.euclidean(Min_BA_Coord, Max_LVA_Coord) > 5:    
            if distance.euclidean(Max_LVA_Coord, Max_RVA_Coord) > 5:
                COORD1.append(Min_BA_Coord)
                COORD2.append(Max_LVA_Coord)
                # Add L/RVA centerpoint to the centerpoints array
                x = (Min_BA_Coord[0]+Max_LVA_Coord[0])/2
                y = (Min_BA_Coord[1]+Max_LVA_Coord[1])/2
                z = (Min_BA_Coord[2]+Max_LVA_Coord[2])/2
                Centerpoint_LVA = np.array([x,y,z])
                if len(Centerpoints) > 0:
                    Centerpoints = np.append(Centerpoints, [Centerpoint_LVA], axis = 0)
                else:
                    Centerpoints = np.array([Centerpoint_LVA])
                print()
                print(f"There's a dicontinuity in the high part LVA:")
                print(f"{[Min_BA_Coord, Max_LVA_Coord]}, the associated centerpoint has been appended to Centerpoints Array")
if Max_RVA_Coord is not None and Max_LVA_Coord is not None:
    if distance.euclidean(Max_LVA_Coord, Max_RVA_Coord) > 5:
          dif = Max_RVA_Coord[2] - Max_LVA_Coord[2]
          if dif > 5:
                COORD1.append(Max_LVA_Coord)
                COORD2.append(Max_LVA_Coord)
                if len(Centerpoints) > 0:
                        Centerpoints = np.append(Centerpoints, [Max_LVA_Coord], axis = 0)
                        print()
                        print(f"There's a dicontinuity in the high part LVA: {Max_LVA_Coord}")
                else:
                    Centerpoints = np.array([Max_LVA_Coord])
                    print()
                    print(f"There's a dicontinuity in the high part LVA: {Max_LVA_Coord}")



# Get the distance form all points of the L/RSA to the obtained coordinate of each vertebral arthery segment.
#Get those (2) endpoints which are closer to the Min_L/RVA_Coord
if len(flat_BA) > 0 and len(flat_RVA) > 0:
    if Max_RVA_Coord is not None and Max_LVA_Coord is not None:
        if distance.euclidean(Min_BA_Coord, Max_RVA_Coord) > 5:
            if distance.euclidean(Max_LVA_Coord, Max_RVA_Coord) > 5:
                COORD1.append(Min_BA_Coord)
                COORD2.append(Max_RVA_Coord)
                # Add L/RVA centerpoint to the centerpoints array
                x = (Min_BA_Coord[0]+Max_RVA_Coord[0])/2
                y = (Min_BA_Coord[1]+Max_RVA_Coord[1])/2
                z = (Min_BA_Coord[2]+Max_RVA_Coord[2])/2
                Centerpoint_RVA = np.array([x,y,z])
                if len(Centerpoints) > 0:
                    Centerpoints = np.append(Centerpoints, [Centerpoint_RVA], axis = 0)
                else:
                    Centerpoints = np.array([Centerpoint_RVA])
                print()
                print(f"There's a dicontinuity in the high part RVA:")
                print(f"{[Min_BA_Coord, Max_RVA_Coord]}, the associated centerpoint has been appended to Centerpoints Array")
elif Max_RVA_Coord is not None and Max_LVA_Coord is not None:
    if distance.euclidean(Max_LVA_Coord, Max_RVA_Coord) > 5 :
        dif = Max_LVA_Coord[2] - Max_RVA_Coord[2]
        if dif > 5:
            COORD1.append(Max_RVA_Coord)
            COORD2.append(Max_RVA_Coord)
            if len(Centerpoints) > 0:
                    Centerpoints = np.append(Centerpoints, [Max_RVA_Coord], axis = 0)
                    print()
                    print(f"There's a dicontinuity in the high part RVA: {Max_RVA_Coord}")
            else:
                Centerpoints = np.array([Max_RVA_Coord])
                print()
                print(f"There's a dicontinuity in the high part RVA: {Max_RVA_Coord}")

#######################################################
### Asses each pair of endpoints to it's vesselType ###
#######################################################

# Search in which CellId we find EP coordinates
c1 = []
c2 = []
CellId1 = [] # [Nº EP , CellID]
CellId2 = []
# COORD1 search
for nep in range(len(COORD1)):
    for idx in range(len(segmentsArrayAff)): # idx is CellID
        for idx2 in range(len(segmentsArrayAff[idx])):
            x = segmentsArrayAff[idx][idx2][0]
            y = segmentsArrayAff[idx][idx2][1]
            z = segmentsArrayAff[idx][idx2][2]
            if distance.euclidean((x,y,z), COORD1[nep]) < 1: 
            #if abs(COORD1[nep][0]-x) < 0.5 and abs(COORD1[nep][1]-y) < 0.5 and abs(COORD1[nep][2]-z) < 0.5:
                c1.append(tuple((nep, idx)))
for item in c1:
    if item not in CellId1:
        CellId1.append(item)
# COORD2 search 
for nep in range(len(COORD2)):
    for idx in range(len(segmentsArrayAff)): # idx is CellID
        for idx2 in range(len(segmentsArrayAff[idx])):
            x = segmentsArrayAff[idx][idx2][0]
            y = segmentsArrayAff[idx][idx2][1]
            z = segmentsArrayAff[idx][idx2][2]
            if distance.euclidean((x,y,z), COORD2[nep]) < 1: 
            # if abs(COORD2[nep][0]-x) < 0.5 and abs(COORD2[nep][1]-y) < 0.5 and abs(COORD2[nep][2]-z) < 0.5:
                c2.append(tuple((nep, idx)))
for item in c2:
    if item not in CellId2:
        CellId2.append(item)

# The main false detections it seems to be appearing are those endpoints which have achived to surpass the condition of the minimum distance due to
# the distance between the endpoints located on the R/LECA segment. So that, all the pair of endpoints where one of them is located in the L/RECA segment must
# not be considered as conflictive. 
CellId1 = np.asarray(CellId1)
CellId2 = np.asarray(CellId2)
for idx in range(len(CellId1)):
    if cellIdToVesselType[CellId1[idx][1]] == 11 or cellIdToVesselType[CellId1[idx][1]] == 12 : # L/RECA Vessel Type
        for idx2 in range(len(CellId2)):
            if CellId2[idx2][0] == CellId1[idx][0]:
                CellId2[idx2][1] = CellId1[idx][1]
for idx in range(len(CellId2)):
    if cellIdToVesselType[CellId2[idx][1]] == 11 or cellIdToVesselType[CellId2[idx][1]] == 12 : # L/RECA Vessel Type
        for idx2 in range(len(CellId1)):
            if CellId1[idx2][0] == CellId2[idx][0]:
                CellId1[idx2][1] = CellId2[idx][1]

# Endpoints locations
if len(CellId1) == len(CellId2):
    for idx in range(len(CellId1)):
        print()
        print(f"The conflictive detected endpoint nº {CellId1[idx][0]} is located between {edgeTypesDict[cellIdToVesselType[CellId1[idx][1]]]}-{edgeTypesDict[cellIdToVesselType[CellId2[idx][1]]]} segments")
else:
    for idx in range(len(CellId1)):
        print()
        print(f"The conflictive detected endpoint COORD1 nº {CellId1[idx][0]} is located in the {edgeTypesDict[cellIdToVesselType[CellId1[idx][1]]]} segments")
    for idx in range(len(CellId2)):
        print()
        print(f"The conflictive detected endpoint COORD2 nº {CellId2[idx][0]} is located in the {edgeTypesDict[cellIdToVesselType[CellId2[idx][1]]]} segments")

# Create an Excel file to save the detected errors
wb = None
wb = openpyxl.Workbook()
w_dis = wb.active
# Set Excel Structure {ID, Distinct Error Locations, Confussion Matrix , Metrics}
w_dis.cell(row = 1, column = 1, value = "Id")
w_dis.cell(row = 2, column = 1, value = caseId[:-7])
w_dis.cell(row = 1, column = 2, value = "LVA low")
w_dis.cell(row = 1, column = 3, value = "LVA mid")
w_dis.cell(row = 1, column = 4, value = "LVA high")
w_dis.cell(row = 1, column = 5, value = "RVA low")
w_dis.cell(row = 1, column = 6, value = "RVA mid")
w_dis.cell(row = 1, column = 7, value = "RVA high")
w_dis.cell(row = 1, column = 8, value = "LCCA / LICA")
w_dis.cell(row = 1, column = 9, value = "RCCA / RICA")
w_dis.cell(row = 4, column = 3, value = "TP")
w_dis.cell(row = 4, column = 4, value = "TN")
w_dis.cell(row = 4, column = 5, value = "FP")
w_dis.cell(row = 4, column = 6, value = "FN")
w_dis.cell(row = 7, column = 2, value = "Precision")
w_dis.cell(row = 7, column = 3, value = "Recall")
w_dis.cell(row = 7, column = 4, value = "F1 score")


Centerpoints_Detected = []        
for idx in range(len(CellId1)):
    if cellIdToVesselType[CellId1[idx][1]] == 7: # RVA
            Centerpoints_Detected.append(Centerpoints[CellId1[idx][0]])
            if COORD1[CellId1[idx][0]][2]-minH > 0.70 * Height: # High AV
                w_dis.cell(row = 2, column = 7, value = 1)
                print("Insert RVA high")
            if COORD1[CellId1[idx][0]][2]-minH > 0.50 * Height and COORD1[CellId1[idx][0]][2]-minH < 0.70 * Height: # Mid AV
                w_dis.cell(row = 2, column = 6, value = 1)
                print("Insert RVA mid")
            if COORD1[CellId1[idx][0]][2]-minH < 0.50 * Height: # Low AV
                w_dis.cell(row = 2, column = 5, value = 1)
                print("Insert RVA low")        
    if cellIdToVesselType[CellId1[idx][1]] == 8: # LVA
            Centerpoints_Detected.append(Centerpoints[CellId1[idx][0]])
            if COORD1[CellId1[idx][0]][2]-minH > 0.70 * Height: # High AV
                w_dis.cell(row = 2, column = 4, value = 1)
                print("Insert LVA high")
            if COORD1[CellId1[idx][0]][2]-minH > 0.50 * Height and COORD1[CellId1[idx][0]][2]-minH < 0.70 * Height: # Mid AV
                w_dis.cell(row = 2, column = 3, value = 1)
                print("Insert LVA mid")
            if COORD1[CellId1[idx][0]][2]-minH < 0.50 * Height: # Low AV
                w_dis.cell(row = 2, column = 2, value = 1)
                print("Insert LVA low")
    if cellIdToVesselType[CellId1[idx][1]] == 10: # LICA
        w_dis.cell(row = 2, column = 8, value = 1)
        Centerpoints_Detected.append(Centerpoints[CellId1[idx][0]])
        print("Insert LICA")
    if cellIdToVesselType[CellId1[idx][1]] == 9:  # RICA
        w_dis.cell(row = 2, column = 9, value = 1)
        Centerpoints_Detected.append(Centerpoints[CellId1[idx][0]])
        print("Insert RICA")
    if cellIdToVesselType[CellId1[idx][1]] == 3:  # RCCA
        w_dis.cell(row = 2, column = 9, value = 1)
        Centerpoints_Detected.append(Centerpoints[CellId1[idx][0]])
        print("Insert RCCA")
    if cellIdToVesselType[CellId1[idx][1]] == 4:  # LCCA
        w_dis.cell(row = 2, column = 8, value = 1)
        Centerpoints_Detected.append(Centerpoints[CellId1[idx][0]])
        print("Insert LCCA")
        
for idx in range(len(CellId2)):
    if cellIdToVesselType[CellId2[idx][1]] == 7: # RVA
            Centerpoints_Detected.append(Centerpoints[CellId2[idx][0]])
            if COORD1[CellId2[idx][0]][2]-minH > 0.70 * Height: # High AV
                w_dis.cell(row = 2, column = 7, value = 1)
                print("Insert RVA high")
            if COORD1[CellId2[idx][0]][2]-minH > 0.50 * Height and COORD1[CellId2[idx][0]][2]-minH < 0.70 * Height: # Mid AV
                w_dis.cell(row = 2, column = 6, value = 1)
                print("Insert RVA mid")
            if COORD1[CellId2[idx][0]][2]-minH < 0.50 * Height: # Low AV
                w_dis.cell(row = 2, column = 5, value = 1)
                print("Insert RVA low")        
    if cellIdToVesselType[CellId2[idx][1]] == 8: # LVA
            Centerpoints_Detected.append(Centerpoints[CellId2[idx][0]])
            if COORD1[CellId2[idx][0]][2]-minH > 0.70 * Height: # High AV
                w_dis.cell(row = 2, column = 4, value = 1)
                print("Insert LVA high")
            if COORD1[CellId2[idx][0]][2]-minH > 0.50 * Height and COORD1[CellId2[idx][0]][2]-minH < 0.70 * Height: # Mid AV
                w_dis.cell(row = 2, column = 3, value = 1)
                print("Insert LVA mid")
            if COORD1[CellId2[idx][0]][2]-minH < 0.50 * Height: # Low AV
                w_dis.cell(row = 2, column = 2, value = 1)
                print("Insert LVA low")
    if cellIdToVesselType[CellId2[idx][1]] == 10: # LICA
        w_dis.cell(row = 2, column = 8, value = 1)
        Centerpoints_Detected.append(Centerpoints[CellId2[idx][0]])
        print("Insert LICA")
    if cellIdToVesselType[CellId2[idx][1]] == 9:  # RICA
        w_dis.cell(row = 2, column = 9, value = 1)
        Centerpoints_Detected.append(Centerpoints[CellId2[idx][0]])
        print("Insert RICA")
    if cellIdToVesselType[CellId2[idx][1]] == 3:  # RCCA
        w_dis.cell(row = 2, column = 9, value = 1)
        Centerpoints_Detected.append(Centerpoints[CellId2[idx][0]])
        print("Insert RCCA")
    if cellIdToVesselType[CellId2[idx][1]] == 4:  # LCCA
        w_dis.cell(row = 2, column = 8, value = 1)
        Centerpoints_Detected.append(Centerpoints[CellId2[idx][0]])
        print("Insert LCCA")
# Eliminate duplicated Centerpoints detections
Final_Centerpoints = []
Memory = []
if len(Centerpoints_Detected) > 1:
    for item in Centerpoints_Detected:
        if item[0] not in Memory:
            Memory.append(item[0])
            Final_Centerpoints.append(item)
else:
    Final_Centerpoints = Centerpoints_Detected

print()        
print(f" We've kept {len(Final_Centerpoints)} centerpoints, it just lasts to check if they are separated enough from each other.")
print()

##############################
### Labeled Excel Filling  ###
##############################
Excel_file = os.path.join(caseDir, f"Labeled_ID_{caseId[:-7]}.xlsx")
wb_label = openpyxl.load_workbook(filename = Excel_file)
# Seleciono la Hoja
w_label = wb_label.active

# Confussion matrix
value_TP = 0
value_TN = 0
value_FP = 0
value_FN = 0

for idx_col in range(2, 10):
    if w_dis.cell(row = 2, column =idx_col).value == 1:
        if w_dis.cell(row = 2, column =idx_col).value == w_label.cell(row = 2, column = idx_col).value: 
            value_TP += 1
        else:
            value_FP += 1
    else: 
        if w_dis.cell(row = 2, column =idx_col).value == w_label.cell(row = 2, column = idx_col).value:
            value_TN += 1
        else:
            value_FN += 1
       
# Include the confusiion matrix in the Excel file to obtain distinct metric avaluations
w_dis.cell(row = 5, column = 3, value = value_TP)
w_dis.cell(row = 5, column = 4, value = value_TN)
w_dis.cell(row = 5, column = 5, value = value_FP)
w_dis.cell(row = 5, column = 6, value = value_FN)

Precision = 0
Recall = 0
F1_score = 0

if (value_TP + value_FP) > 0:
    Precision = value_TP / (value_TP + value_FP)
    w_dis.cell(row = 8, column = 2, value = Precision)


if (value_TP + value_FN) > 0:
    Recall = value_TP / (value_TP + value_FN)
    w_dis.cell(row = 8, column = 3, value = Recall)


if (Precision+Recall) > 0:
    F1_score = 2 * Precision * Recall / (Precision+Recall)
    w_dis.cell(row = 8, column = 4, value = F1_score)




# Save Detected.xlsx in it's case Directory
wb.save(os.path.join(caseDir,"Pred_avaluation.xlsx"))

###################################
### Close Centerpoints Checking ###
###################################

def DetectCloseCenterpoints(Final_Centerpoints):
    # Final_Centerpoints is a list containing the centerpoints coordinates as [array([x1, y1, z1], array(x2, y2, z2)etc.)
    if len(Final_Centerpoints) > 1:
        IndexX = []
        IndexY = np.empty(shape = (len(Final_Centerpoints), len(Final_Centerpoints)))
        for idx in range(len(Final_Centerpoints)):
            for idx2 in range(len(Final_Centerpoints)):
                Dist = distance.euclidean(Final_Centerpoints[idx], Final_Centerpoints[idx2])
                IndexY[idx][0] = int(idx)
                if  Dist < 15 and Dist > 0 and idx2 not in IndexX:
                    IndexY[idx][idx2] = idx2
                    IndexX.append(idx2) # Index of centerpoints to be deleted 
                else: 
                    IndexY[idx][idx2] = None

        IndexX.sort()
        IndexY = np.array(IndexY)
        Add_centerpoints = []

        for idx in range(len(IndexY)):
            IndexY[idx].sort()
            CP = []
            for idx2 in range(len(IndexY[idx])):
                if math.isnan(IndexY[idx][idx2]) ==  False:
                    CP.append(Final_Centerpoints[int(IndexY[idx][idx2])])
            Put_CP = sum(CP)/len(CP)
            if len(Add_centerpoints) >= 1:
                if distance.euclidean(Put_CP, Add_centerpoints[-1]) > 10:        
                        Add_centerpoints.append(Put_CP)
            else: 
                Add_centerpoints.append(Put_CP)

        Def_Add_Centerpoints = []

        if len(Add_centerpoints) > 1: 
            for idx in range(len(Add_centerpoints)):
                for idx2 in range(len(Add_centerpoints)):
                    D1 = distance.euclidean(Add_centerpoints[idx2], Add_centerpoints[idx])
                    if   D1 > 15 :
                        Def_Add_Centerpoints.append(Add_centerpoints[idx2])
        else:
            Def_Add_Centerpoints = Add_centerpoints
        Def_Centerpoints = []
        for idx, num in enumerate(Final_Centerpoints):
            if idx not in IndexX:
                Def_Centerpoints.append(num)
        for idx in Def_Add_Centerpoints: 
            Def_Centerpoints.append(idx)
        Centerpoints_RAS = []
        Memory = []
        for idx in range(len(Def_Centerpoints)):
             if sum(Def_Centerpoints[idx]) not in Memory:
                    Memory.append(sum(Def_Centerpoints[idx]))
                    Centerpoints_RAS.append(Def_Centerpoints[idx])
        print(f"Finally, from the initial {len(Final_Centerpoints)} centerpoints we obtain {len(Centerpoints_RAS)} in RAS coordinate/s")
    else:
        Centerpoints_RAS = Final_Centerpoints
        print("As the lenght of Centerpoints array it's lower than 2, there's no close endpoints, so that we proceed to extract the 3D patches")
    return Centerpoints_RAS

Centerpoints_RAS = None
print("Close Centerpoints - Iteration_1:")
Centerpoints_RAS_prev = DetectCloseCenterpoints(Final_Centerpoints)
if len(Centerpoints_RAS_prev) > 1: 
    print()
    print("Check for new Close Centerpoints - Iteration_2:")
    Centerpoints_RAS = DetectCloseCenterpoints(Centerpoints_RAS_prev)
else:
    Centerpoints_RAS = Centerpoints_RAS_prev
print()
print("Proceed with patch extraction...")

# CTA Patch extraction

caseId_CTA = caseId[:-7] + "_CTA.nii.gz"
aff_CTA = nib.load(os.path.join(caseDir, caseId_CTA))
slicer.util.loadLabelVolume(os.path.join(caseDir, caseId_CTA))

#print(aff_CTA)
CTANode = None
CTANode = getNode(caseId_CTA[:-7]) # caseId_CTA format XXXXXXXX_CTA.nii.gz
# Obtain NParray
CTAarray = slicer.util.arrayFromVolume(CTANode)
# Apply affine to centerpoints coordinates
if len(Centerpoints_RAS) < 1:
    print("No conflictive endpoints detected")
else:
    Centerpoints_ijk = None
    Centerpoints_ijk = apply_affine(np.linalg.inv(aff), Centerpoints_RAS)
    # Cretae a [48 x 48 x 48] CTA patch for each detection
    for idx in range(len(Centerpoints_ijk)):
            A = np.asarray(CTAarray[int(Centerpoints_ijk[idx][2]) -32 : int(Centerpoints_ijk[idx][2]) + 32, 
                                    int(Centerpoints_ijk[idx][1]) -32 : int(Centerpoints_ijk[idx][1]) + 32, 
                                    int(Centerpoints_ijk[idx][0]) -32 : int(Centerpoints_ijk[idx][0]) + 32])
            ni_img = nib.Nifti1Image(A, aff)
            nib.save(ni_img, os.path.join(caseDir, "patch_volume_CTA_{0}_{1}.nii.gz".format(caseId_CTA[:-11], idx)))
            print(f'Patch volume CTA {caseId_CTA[:-11]}_{idx} has been saved')

# Label Patch extraction

caseId_label = caseId[:-7] + "_lab.nii.gz"
aff_label = nib.load(os.path.join(caseDir, caseId_label))
slicer.util.loadLabelVolume(os.path.join(caseDir, caseId_label))

LabelNode = None
LabelNode = getNode(caseId_label[:-7])
Labelarray = slicer.util.arrayFromVolume(LabelNode)
if len(Centerpoints_RAS) < 1:
    print("No conflictive endpoints detected")
else:
    Centerpoints_ijk = None
    Centerpoints_ijk = apply_affine(np.linalg.inv(aff), Centerpoints_RAS)
    # Cretae a [48 x 48 x 48] CTA patch for each detection
    for idx in range(len(Centerpoints_ijk)):
            B = np.asarray(Labelarray[int(Centerpoints_ijk[idx][2]) -32 : int(Centerpoints_ijk[idx][2]) + 32, 
                                    int(Centerpoints_ijk[idx][1]) -32 : int(Centerpoints_ijk[idx][1]) + 32, 
                                    int(Centerpoints_ijk[idx][0]) -32 : int(Centerpoints_ijk[idx][0]) + 32])
            ni_img = nib.Nifti1Image(B, aff)
            nib.save(ni_img, os.path.join(caseDir, "patch_volume_Label_{0}_{1}.nii.gz".format(caseId_label[:-11], idx)))
            print(f'Patch volume Label {caseId_label[:-11]}_{idx} has been saved')

# Save centerpoints in a .npy format
newpath = os.path.join(caseDir, "Centerpoints")
if not os.path.exists(newpath):
    os.makedirs(newpath)
for idx in range(len(Centerpoints_RAS)): 
    CP_save = np.array(Centerpoints[idx])
    np.save((os.path.join(newpath, f"Centerpoints_{idx}.npy")), CP_save)

